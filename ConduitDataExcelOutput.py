import os, json

from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook

path = 'D:/Data/conduit_data.json'

with open(path) as f:
	conduits = json.load(f)

	wb = Workbook()

	ws1 = wb.create_sheet('Summary')
	ws2 = wb.create_sheet('Receiver Breakdown')
	ws3 = wb.create_sheet('Contributor Breakdown')
	ws4 = wb.create_sheet('Totals By Receiver')
	ws5 = wb.create_sheet('Totals By Contributor')

	ws1['A1'] = 'Conduit Name'
	ws1['B1'] = 'Address'
	ws1['C1'] = 'Email'
	ws1['D1'] = '2020 Cycle'
	ws1['E1'] = '2022 Cycle'
	ws1['F1'] = '2024 Cycle'

	ws2['A1'] = 'Conduit Name'
	ws2['B1'] = 'Receiver'
	ws2['C1'] = 'Address'
	ws2['D1'] = '2020 Cycle'
	ws2['E1'] = '2022 Cycle'
	ws2['F1'] = '2024 Cycle'

	ws3['A1'] = 'Conduit Name'
	ws3['B1'] = 'Contributor'
	ws3['C1'] = 'Address'
	ws3['D1'] = '2020 Cycle'
	ws3['E1'] = '2022 Cycle'
	ws3['F1'] = '2024 Cycle'

	ws4['A1'] = 'Receiver Name'
	ws4['B1'] = 'Address'
	ws4['C1'] = '2020 Cycle'
	ws4['D1'] = '2022 Cycle'
	ws4['E1'] = '2024 Cycle'

	ws5['A1'] = 'Contributor Name'
	ws5['B1'] = 'Address'
	ws5['C1'] = '2020 Cycle'
	ws5['D1'] = '2022 Cycle'
	ws5['E1'] = '2024 Cycle'

	greystyle = Font(color="888888", italic=True)

	x = 2
	for key in conduits:
		conduit = conduits[key]
		ws1.cell(row=x, column=1).value = conduit['name']
		ws1.cell(row=x, column=2).value = conduit['address'].strip()
		ws1.cell(row=x, column=3).value = conduit['email']
		ws1.cell(row=x, column=4).value = conduit['2020_cycle_total']
		ws1.cell(row=x, column=5).value = conduit['2022_cycle_total']
		ws1.cell(row=x, column=6).value = conduit['2024_cycle_total']
		x += 1

	x = 2
	for key in conduits:
		conduit = conduits[key]
		ws2.cell(row=x, column=1).value = conduit['name']
		ws2.cell(row=x, column=2).value = '(Total)'
		ws2.cell(row=x, column=2).font = greystyle
		ws2.cell(row=x, column=4).value = conduit['2020_cycle_total']
		ws2.cell(row=x, column=5).value = conduit['2022_cycle_total']
		ws2.cell(row=x, column=6).value = conduit['2024_cycle_total']
		x += 1

		receivers = {}
		for entry in conduit['transfers']:
			if not entry['receiver_name'] in receivers:
				receiver = {}
				receiver['name'] = entry['receiver_name']
				receiver['address'] = entry['address']
				receiver['2020_cycle_total'] = 0
				receiver['2022_cycle_total'] = 0
				receiver['2024_cycle_total'] = 0
				receivers[entry['receiver_name']] = receiver
			year = int(entry['date'].split('/')[-1])
			if year == 2019 or year == 2020:
				receivers[entry['receiver_name']]['2020_cycle_total'] += entry['amount']
			elif year == 2021 or year == 2022:
				receivers[entry['receiver_name']]['2022_cycle_total'] += entry['amount']
			elif year == 2023 or year == 2024:
				receivers[entry['receiver_name']]['2024_cycle_total'] += entry['amount']

		for receiver in receivers.values():
			ws2.cell(row=x, column=1).value = conduit['name']
			ws2.cell(row=x, column=1).font = greystyle
			ws2.cell(row=x, column=2).value = receiver['name']
			ws2.cell(row=x, column=3).value = receiver['address'].strip()
			ws2.cell(row=x, column=4).value = receiver['2020_cycle_total']
			ws2.cell(row=x, column=5).value = receiver['2022_cycle_total']
			ws2.cell(row=x, column=6).value = receiver['2024_cycle_total']
			x += 1

	x = 2
	for key in conduits:
		conduit = conduits[key]
		ws3.cell(row=x, column=1).value = conduit['name']
		ws3.cell(row=x, column=2).value = '(Total)'
		ws3.cell(row=x, column=2).font = greystyle
		ws3.cell(row=x, column=4).value = conduit['2020_cycle_total']
		ws3.cell(row=x, column=5).value = conduit['2022_cycle_total']
		ws3.cell(row=x, column=6).value = conduit['2024_cycle_total']
		x += 1

		donors = {}
		for entry in conduit['donations']:
			if not entry['donor_name'] in donors:
				donor = {}
				donor['name'] = entry['donor_name']
				donor['address'] = entry['address']
				donor['2020_cycle_total'] = 0
				donor['2022_cycle_total'] = 0
				donor['2024_cycle_total'] = 0
				donors[entry['donor_name']] = donor
			year = int(entry['date'].split('/')[-1])
			if year == 2019 or year == 2020:
				donors[entry['donor_name']]['2020_cycle_total'] += entry['amount']
			elif year == 2021 or year == 2022:
				donors[entry['donor_name']]['2022_cycle_total'] += entry['amount']
			elif year == 2023 or year == 2024:
				donors[entry['donor_name']]['2024_cycle_total'] += entry['amount']

		for donor in donors.values():
			ws3.cell(row=x, column=1).value = conduit['name']
			ws3.cell(row=x, column=1).font = greystyle
			ws3.cell(row=x, column=2).value = donor['name']
			ws3.cell(row=x, column=3).value = donor['address'].strip()
			ws3.cell(row=x, column=4).value = donor['2020_cycle_total']
			ws3.cell(row=x, column=5).value = donor['2022_cycle_total']
			ws3.cell(row=x, column=6).value = donor['2024_cycle_total']
			x += 1



	wb.save('conduit_data.xlsx')