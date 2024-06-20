from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

filename = "C:/Users/choco/Downloads/game_data.xlsx"

wb = load_workbook(filename)
sheet = wb['game_data']

characters = []
for row in range(3,79):
    characters.append(sheet.cell(row=row, column=15).value)

def getMatchups(character):
    index = characters.index(character)
    result = []
    for i in range(76):
        result.append(sheet.cell(row=3+index,column=16+i).value)
    return result    

def compareMatchups(characterA, characterB):
    a = getMatchups(characterA)
    b = getMatchups(characterB)
    x = 0
    for i in range(len(a)):
        x += abs(a[i] - b[i])
    return x / len(characters)

print(compareMatchups("Fox","Fox"))

for a in range(76):
    for b in range(76):
        cell = sheet.cell(row=3+a+76+2, column=16+b)
        cell.value = compareMatchups(characters[a],characters[b])
        
wb.save(filename)
